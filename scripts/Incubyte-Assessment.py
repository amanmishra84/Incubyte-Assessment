import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook

# Define the SQL Server connection string
server = 'server_name'
database = 'Incubyte'
connection_string = f'mssql+pyodbc://@{server}/{database}?driver=ODBC+Driver+17+for+SQL+Server&trusted_connection=yes'

# Create a SQLAlchemy engine
engine = create_engine(connection_string)

# Load the Excel file
excel_file_path = 'data/assessment_dataset.xlsx'
wb = load_workbook(excel_file_path)
ws = wb.active

# Determine the column names from the first row
columns = [cell.value for cell in next(ws.iter_rows(max_row=1))]

# Define chunk size
chunksize = 10000  # Adjust chunk size based on your system's memory capacity

# Process data in chunks and write to SQL Server
table_name = 'sales_data'
batch = []

for row in ws.iter_rows(min_row=2, values_only=True):
    batch.append(row)
    if len(batch) == chunksize:
        df_chunk = pd.DataFrame(batch, columns=columns)
        df_chunk.to_sql(table_name, engine, if_exists='append', index=False)
        batch = []

# Write any remaining rows
if batch:
    df_chunk = pd.DataFrame(batch, columns=columns)
    df_chunk.to_sql(table_name, engine, if_exists='append', index=False)

print("Data has been written to the SQL Server database successfully.")
